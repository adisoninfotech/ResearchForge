"""Content extraction for supported file kinds."""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from typing import Any

from app.models.enums import FileKind


@dataclass
class ExtractedPage:
    page_number: int
    text: str


@dataclass
class ExtractionResult:
    title: str | None
    plain_text: str
    pages: list[ExtractedPage] = field(default_factory=list)
    metadata: dict[str, Any] = field(default_factory=dict)
    bib_entries: list[dict[str, Any]] = field(default_factory=list)


def extract_content(*, kind: FileKind, data: bytes, filename: str) -> ExtractionResult:
    if kind == FileKind.PDF:
        return _extract_pdf(data)
    if kind == FileKind.DOCX:
        return _extract_docx(data)
    if kind == FileKind.XLSX:
        return _extract_xlsx(data)
    if kind == FileKind.CSV:
        return _extract_csv(data)
    if kind == FileKind.BIBTEX:
        return _extract_bibtex(data)
    if kind == FileKind.RIS:
        return _extract_ris(data)
    if kind in {FileKind.TXT, FileKind.MARKDOWN}:
        text = data.decode("utf-8", errors="replace")
        return ExtractionResult(title=filename, plain_text=text, pages=[ExtractedPage(1, text)])
    if kind in {FileKind.PNG, FileKind.JPEG}:
        return ExtractionResult(
            title=filename,
            plain_text="",
            pages=[],
            metadata={"figure": True, "note": "Figure uploaded; OCR not enabled"},
        )
    raise ValueError("Unsupported file kind for extraction")


def _extract_pdf(data: bytes) -> ExtractionResult:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover
        raise ValueError("PDF support unavailable") from exc
    reader = PdfReader(io.BytesIO(data))
    pages: list[ExtractedPage] = []
    parts: list[str] = []
    for idx, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        pages.append(ExtractedPage(idx, text))
        parts.append(text)
    title = None
    if reader.metadata and reader.metadata.title:
        title = str(reader.metadata.title)
    return ExtractionResult(title=title, plain_text="\n\n".join(parts), pages=pages)


def _extract_docx(data: bytes) -> ExtractionResult:
    try:
        from docx import Document
    except ImportError as exc:  # pragma: no cover
        raise ValueError("DOCX support unavailable") from exc
    doc = Document(io.BytesIO(data))
    paras = [p.text for p in doc.paragraphs if p.text.strip()]
    text = "\n\n".join(paras)
    return ExtractionResult(title=None, plain_text=text, pages=[ExtractedPage(1, text)])


def _extract_xlsx(data: bytes) -> ExtractionResult:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ValueError("XLSX support unavailable") from exc
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows_out: list[str] = []
    for sheet in wb.worksheets:
        rows_out.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            if any(cells):
                rows_out.append("\t".join(cells))
    text = "\n".join(rows_out)
    return ExtractionResult(title=None, plain_text=text, pages=[ExtractedPage(1, text)])


def _extract_csv(data: bytes) -> ExtractionResult:
    text = data.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    lines = [", ".join(row) for row in reader]
    body = "\n".join(lines)
    return ExtractionResult(title=None, plain_text=body, pages=[ExtractedPage(1, body)])


def _extract_bibtex(data: bytes) -> ExtractionResult:
    text = data.decode("utf-8", errors="replace")
    entries = parse_bibtex(text)
    summaries = []
    for entry in entries:
        title = entry.get("title") or "(missing title)"
        summaries.append(f"{entry.get('cite_key', '?')}: {title}")
    return ExtractionResult(
        title="BibTeX import",
        plain_text="\n".join(summaries),
        pages=[ExtractedPage(1, text)],
        bib_entries=entries,
        metadata={"format": "bibtex", "count": len(entries)},
    )


def _extract_ris(data: bytes) -> ExtractionResult:
    text = data.decode("utf-8", errors="replace")
    entries = parse_ris(text)
    summaries = [e.get("title") or "(missing title)" for e in entries]
    return ExtractionResult(
        title="RIS import",
        plain_text="\n".join(summaries),
        pages=[ExtractedPage(1, text)],
        bib_entries=entries,
        metadata={"format": "ris", "count": len(entries)},
    )


def parse_bibtex(text: str) -> list[dict[str, Any]]:
    """Minimal BibTeX parser — does not invent missing fields."""
    entries: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("%"):
            continue
        if line.startswith("@"):
            if current:
                entries.append(current)
            # @article{key,
            kind_key = line[1:].split("{", 1)
            entry_type = kind_key[0].strip().lower()
            cite_key = ""
            if len(kind_key) > 1:
                cite_key = kind_key[1].rstrip(",").strip()
            current = {"entry_type": entry_type, "cite_key": cite_key, "authors": []}
            continue
        if current is None:
            continue
        if line.startswith("}"):
            entries.append(current)
            current = None
            continue
        if "=" in line:
            key, value = line.split("=", 1)
            key = key.strip().lower()
            value = value.strip().rstrip(",").strip()
            value = value.strip("{}").strip('"').strip()
            if key == "author":
                parts = value.replace(" and ", "|").split("|")
                current["authors"] = [a.strip() for a in parts if a.strip()]
            elif key == "year":
                try:
                    current["year"] = int("".join(ch for ch in value if ch.isdigit())[:4])
                except ValueError:
                    pass
            else:
                current[key] = value
    if current:
        entries.append(current)
    return entries


def parse_ris(text: str) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    current: dict[str, Any] = {"authors": []}
    for raw in text.splitlines():
        line = raw.rstrip()
        if not line:
            continue
        if line.startswith("TY  -"):
            if current.get("title") or current.get("authors"):
                entries.append(current)
            current = {"authors": [], "entry_type": line[5:].strip()}
            continue
        if line.startswith("ER  -"):
            entries.append(current)
            current = {"authors": []}
            continue
        if "  - " not in line:
            continue
        tag, value = line.split("  - ", 1)
        tag = tag.strip().upper()
        value = value.strip()
        if tag == "TI":
            current["title"] = value
        elif tag == "AU":
            current.setdefault("authors", []).append(value)
        elif tag == "PY":
            try:
                current["year"] = int(value[:4])
            except ValueError:
                pass
        elif tag == "JO":
            current["venue"] = value
        elif tag == "DO":
            current["doi"] = value
        elif tag == "UR":
            current["url"] = value
        elif tag == "AB":
            current["abstract"] = value
    if current.get("title") or current.get("authors"):
        entries.append(current)
    return entries
